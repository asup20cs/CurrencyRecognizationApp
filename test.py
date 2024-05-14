import openpyxl
import os
from tensorflow.keras.models import load_model
from PIL import Image
import numpy as np
# Create a new workbook
workbook = openpyxl.Workbook()

# Select the active sheet
sheet = workbook.active

# Set the column headers
sheet['A1'] = 'File name'
sheet['B1'] = 'Actual denomination'
sheet['C1'] = 'Predicted denomination'
sheet['D1'] = 'Correct prediction ?'

# Save the workbook
workbook.save('D:/MajorProject/App/test.xlsx')

folder_path = "C:/Users/ashut/Downloads/test"

# Get the list of image files in the folder
image_files = [file for file in os.listdir(folder_path) if file.endswith('.jpg')]

# Iterate through the image files
for i, image_file in enumerate(image_files):
    # Extract the denomination from the file name
    denomination = image_file.split('_')[0]

    # Write the denomination in the "Actual denomination" column
    sheet.cell(row=i+2, column=2).value = int(denomination)

    #write the filenames form the folder to the excel sheet
for i, image_file in enumerate(image_files):
    # Write the file name in the "File name" column
    sheet.cell(row=i+2, column=1).value = image_file

# Save the workbook again
workbook.save('D:/MajorProject/App/test.xlsx')

#now we will predict the denominations of the images using a model and write the predicted denominations that are returned by the model for each image in the excel sheet
model = load_model("models/indian.h5")
labels = ["10", "100", "20", "200", "2000", "50", "500"]

# Iterate through the image files
for i, image_file in enumerate(image_files):
    # Open the image
    image_path = os.path.join(folder_path, image_file)
    img = Image.open(image_path)
    img = img.resize((240, 240))  # Adjust size based on your model
    x = np.array(img)
    x = x / 255.0
    x = np.expand_dims(x, axis=0)

    # Make prediction
    prediction = model.predict(x)

    # Interpret the prediction
    predicted_class_index = np.argmax(prediction[0])
    predicted_currency = labels[predicted_class_index]

    # Write the predicted denomination in the "Predicted denomination" column
    sheet.cell(row=i+2, column=3).value = int(predicted_currency)

# Save the workbook again
workbook.save('D:/MajorProject/App/test.xlsx')

